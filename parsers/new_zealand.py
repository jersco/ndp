from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common.cli import resolve_output_path
from common.contracts import FieldKind, FieldSpec
from common.output import write_jsonl
from common.units import convert_numeric_value, extract_unit_from_label
from nutrient_mapping import CORE_FIELD_UNITS, NEW_ZEALAND_FIELD_SPECS

DEFAULT_SHEET_NAME = "Concisen Tables 14th Edition wi"
SOURCE_NAME = "new_zealand"

FOOD_ID_PATTERN = re.compile(r"^[A-Z]\d+$")
CATEGORY_ID_PATTERN = re.compile(r"^[A-Z]$")
PORTION_NAME_PATTERN = re.compile(
    r"^(?:\d|½|¼|¾|one\b|two\b|three\b|four\b|five\b|single\b|double\b|half\b|quarter\b|cup\b|tbsp\b|tsp\b|slice\b|serving\b)",
    re.IGNORECASE,
)


def normalize_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        candidate = value.replace(",", "").strip()
        if not candidate:
            return None
        if candidate.casefold() == "trace":
            return 0.0
        try:
            return float(candidate)
        except ValueError:
            return value.strip()
    return value


def canonicalize_header(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.strip().split()).casefold()


def specs_to_source_map(specs: dict[str, FieldSpec]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field, spec in specs.items():
        if spec.kind in {FieldKind.SOURCE, FieldKind.LITERAL} and isinstance(spec.source, str):
            mapping[field] = spec.source
        else:
            mapping[field] = ""
    return mapping


def resolve_source_columns(
    headers: list[Any], mapping: dict[str, str]
) -> dict[str, str | None]:
    header_values = [h for h in headers if isinstance(h, str)]
    headers_by_canonical = {canonicalize_header(h): h for h in header_values}

    resolved: dict[str, str | None] = {}
    for target_field, source_column in mapping.items():
        if source_column == "" or target_field in {"source", "portions"}:
            resolved[target_field] = source_column
            continue

        if source_column in header_values:
            resolved[target_field] = source_column
            continue

        source_canonical = canonicalize_header(source_column)
        direct_match = headers_by_canonical.get(source_canonical)
        if direct_match is not None:
            resolved[target_field] = direct_match
            continue

        startswith_match = next(
            (
                header
                for header in header_values
                if canonicalize_header(header).startswith(source_canonical)
            ),
            None,
        )
        resolved[target_field] = startswith_match

    return resolved


def resolve_measure_column(headers: list[Any]) -> str | None:
    for header in headers:
        if canonicalize_header(header) == "measure":
            return header
    return None


def build_unit_by_header(headers: list[Any], units_row: list[Any]) -> dict[str, str | None]:
    units_by_header: dict[str, str | None] = {}
    for index, header in enumerate(headers):
        if not isinstance(header, str):
            continue
        if index >= len(units_row):
            units_by_header[header] = None
            continue
        unit_value = units_row[index]
        if isinstance(unit_value, str):
            candidate = unit_value.strip()
            units_by_header[header] = candidate or None
        else:
            units_by_header[header] = None
    return units_by_header


def build_source_units(
    resolved_mapping: dict[str, str | None], units_by_header: dict[str, str | None]
) -> dict[str, str | None]:
    source_units: dict[str, str | None] = {}
    for target_field, source_column in resolved_mapping.items():
        if target_field in {"source", "portions"}:
            continue
        if source_column in ("", None):
            continue
        unit = units_by_header.get(source_column)
        if unit is None:
            unit = extract_unit_from_label(source_column)
        source_units[target_field] = unit
    return source_units


def is_food_id(value: Any) -> bool:
    return isinstance(value, str) and FOOD_ID_PATTERN.fullmatch(value.strip()) is not None


def is_category_id(value: Any) -> bool:
    return isinstance(value, str) and CATEGORY_ID_PATTERN.fullmatch(value.strip()) is not None


def is_hundred_gram_measure(value: Any) -> bool:
    return isinstance(value, (int, float)) and abs(float(value) - 100.0) < 1e-9


def looks_like_portion_name(value: Any) -> bool:
    return isinstance(value, str) and PORTION_NAME_PATTERN.match(value.strip()) is not None


def build_portion_entry(name: Any, weight: Any) -> dict[str, Any] | None:
    if not isinstance(weight, (int, float)):
        return None
    numeric_weight = float(weight)
    if numeric_weight <= 0:
        return None
    normalized_name = name.strip() if isinstance(name, str) else str(name)
    return {
        "name": normalized_name,
        "amount": numeric_weight,
        "unit": "g",
    }


def map_new_zealand_row(
    raw_row: dict[str, Any],
    resolved_mapping: dict[str, str | None],
    source_units: dict[str, str | None],
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    for target_field, source_column in resolved_mapping.items():
        if target_field == "source":
            normalized[target_field] = source_column
            continue

        if target_field == "portions":
            normalized[target_field] = []
            continue

        if source_column in ("", None):
            normalized[target_field] = None
            continue

        value = normalize_value(raw_row.get(source_column))
        if isinstance(value, (int, float)):
            value = convert_numeric_value(
                value,
                source_units.get(target_field),
                CORE_FIELD_UNITS.get(target_field),
                field=target_field,
            )

        normalized[target_field] = value

    return normalized


def finalize_food(food: dict[str, Any]) -> dict[str, Any]:
    portions = food.get("portions") or []
    deduped_portions: list[dict[str, Any]] = []
    seen_portions: set[tuple[str, float]] = set()

    for portion in portions:
        if not isinstance(portion, dict):
            continue
        name = portion.get("name")
        amount = portion.get("amount")
        if not isinstance(name, str) or not isinstance(amount, (int, float)):
            continue
        portion_key = (name, float(amount))
        if portion_key in seen_portions:
            continue
        deduped_portions.append(portion)
        seen_portions.add(portion_key)

    food["portions"] = deduped_portions or None
    return food


def iter_rows(
    workbook_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    field_specs: dict[str, FieldSpec] = NEW_ZEALAND_FIELD_SPECS,
):
    mapping = specs_to_source_map(field_specs)
    wb = load_workbook(filename=workbook_path, data_only=True)
    ws = wb[sheet_name]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    units_row = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
    resolved_mapping = resolve_source_columns(headers, mapping)
    units_by_header = build_unit_by_header(headers, units_row)
    source_units = build_source_units(resolved_mapping, units_by_header)
    measure_column = resolve_measure_column(headers)
    source_id_column = resolved_mapping.get("source_id")
    name_column = resolved_mapping.get("name")

    if source_id_column is None:
        raise RuntimeError("Could not resolve source_id column from mapping")

    current_food: dict[str, Any] | None = None

    for row in ws.iter_rows(min_row=5, values_only=True):
        raw_row = dict(zip(headers, row))
        source_id_raw = raw_row.get(source_id_column)
        source_id = source_id_raw.strip() if isinstance(source_id_raw, str) else source_id_raw
        name_value = raw_row.get(name_column) if name_column is not None else None
        measure_value = (
            normalize_value(raw_row.get(measure_column)) if measure_column is not None else None
        )

        if is_category_id(source_id):
            continue

        if is_food_id(source_id):
            if current_food is not None and current_food.get("source_id") != source_id:
                yield finalize_food(current_food)
                current_food = None

            if current_food is None or is_hundred_gram_measure(measure_value):
                current_food = map_new_zealand_row(raw_row, resolved_mapping, source_units)
                continue

            if (
                isinstance(measure_value, (int, float))
                and not is_hundred_gram_measure(measure_value)
                and looks_like_portion_name(name_value)
            ):
                portion_entry = build_portion_entry(name_value, measure_value)
                if portion_entry is not None:
                    current_food["portions"].append(portion_entry)
            continue

        if current_food is None:
            continue

        if is_hundred_gram_measure(measure_value) and not looks_like_portion_name(name_value):
            yield finalize_food(current_food)
            current_food = None
            continue

        if (
            isinstance(measure_value, (int, float))
            and not is_hundred_gram_measure(measure_value)
            and looks_like_portion_name(name_value)
        ):
            portion_entry = build_portion_entry(name_value, measure_value)
            if portion_entry is not None:
                current_food["portions"].append(portion_entry)

    if current_food is not None:
        yield finalize_food(current_food)


def register_subparser(subparsers):
    parser = subparsers.add_parser("new-zealand", help="Parse New Zealand food composition workbook")
    parser.add_argument("--workbook", required=True, help="Path to New Zealand workbook (.xlsx)")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Workbook sheet name")
    parser.add_argument(
        "--output",
        default=None,
        help="Output JSONL path (default: data/outputs/new_zealand.jsonl, use '-' for stdout)",
    )
    parser.set_defaults(handler=run_from_args)


def run_from_args(args) -> None:
    output_path = resolve_output_path(args.output, SOURCE_NAME)
    rows = iter_rows(Path(args.workbook), sheet_name=args.sheet)
    write_jsonl(rows, output_path)


def main(argv: list[str] | None = None):
    parser = argparse.ArgumentParser(description="New Zealand food composition parser")
    parser.add_argument("--workbook", required=True, help="Path to New Zealand workbook (.xlsx)")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Workbook sheet name")
    parser.add_argument(
        "--output",
        default=None,
        help="Output JSONL path (default: data/outputs/new_zealand.jsonl, use '-' for stdout)",
    )
    args = parser.parse_args(argv)
    run_from_args(args)


if __name__ == "__main__":
    main()
