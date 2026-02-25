from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common.cli import resolve_output_path
from common.contracts import FieldKind, FieldSpec
from common.output import write_jsonl
from common.units import convert_numeric_value, extract_unit_from_label
from nutrient_mapping import COFID_FIELD_SPECS, CORE_FIELD_UNITS, CORE_FOOD_FIELDS

SOURCE_NAME = "cofid"

COFID_SHEETS = {
    "proximates": "1.3 Proximates",
    "inorganics": "1.4 Inorganics",
    "vitamins": "1.5 Vitamins",
    "sfa_food": "1.8 (SFA per 100gFood)",
    "mufa_food": "1.10 (MUFA per 100gFood)",
    "pufa_food": "1.12 (PUFA per 100gFood)",
}

DIRECT_FIELD_MAP: dict[str, tuple[str, str]] = {
    "source_id": ("proximates", "Food Code"),
    "name": ("proximates", "Food Name"),
    "calories": ("proximates", "Energy (kcal)"),
    "protein": ("proximates", "Protein (g)"),
    "dietary_fat": ("proximates", "Fat (g)"),
    "carbohydrates": ("proximates", "Carbohydrate (g)"),
    "net_carbohydrates": ("proximates", "Carbohydrate (g)"),
    "fiber": ("proximates", "AOAC fibre (g)"),
    "starch": ("proximates", "Starch (g)"),
    "sugars": ("proximates", "Total sugars (g)"),
    "saturated_fats": ("proximates", "Satd FA /100g fd (g)"),
    "monounsaturated_fat": ("proximates", "Mono FA /100g food (g)"),
    "polyunsaturated_fat": ("proximates", "Poly FA /100g food (g)"),
    "omega_3_fats": ("proximates", "n-3 poly /100g food (g)"),
    "omega_6_fats": ("proximates", "n-6 poly /100g food (g)"),
    "trans_fats": ("proximates", "Trans FAs /100g food (g)"),
    "dietary_cholesterol": ("proximates", "Cholesterol (mg)"),
    "alcohol": ("proximates", "Alcohol (g)"),
    "water": ("proximates", "Water (g)"),
    "calcium": ("inorganics", "Calcium (mg)"),
    "copper": ("inorganics", "Copper (mg)"),
    "iron": ("inorganics", "Iron (mg)"),
    "manganese": ("inorganics", "Manganese (mg)"),
    "magnesium": ("inorganics", "Magnesium (mg)"),
    "phosphorus": ("inorganics", "Phosphorus (mg)"),
    "potassium": ("inorganics", "Potassium (mg)"),
    "selenium": ("inorganics", "Selenium"),
    "sodium": ("inorganics", "Sodium (mg)"),
    "zinc": ("inorganics", "Zinc (mg)"),
    "vitamin_a_retinol": ("vitamins", "Retinol"),
    "vitamin_b1_thiamin": ("vitamins", "Thiamin (mg)"),
    "vitamin_b2_riboflavin": ("vitamins", "Riboflavin (mg)"),
    "vitamin_b3_niacin": ("vitamins", "Niacin (mg)"),
    "vitamin_b5_pantothenic_acid": ("vitamins", "Pantothenate"),
    "vitamin_b6_pyridoxine": ("vitamins", "Vitamin B6 (mg)"),
    "vitamin_b12_cobalamin": ("vitamins", "Vitamin B12"),
    "folate_vitamin_b9": ("vitamins", "Folate"),
    "vitamin_c_ascorbic_acid": ("vitamins", "Vitamin C (mg)"),
    "vitamin_d_calciferol": ("vitamins", "Vitamin D"),
    "vitamin_e_tocopherol": ("vitamins", "Vitamin E (mg)"),
    "vitamin_k_phylloquinone_and_menaquinone": ("vitamins", "Vitamin K1"),
    "omega_3_ala": ("pufa_food", "cis n-3 C18:3 /100g food (g)"),
    "omega_3_epa": ("pufa_food", "cis n-3 C20:5 /100g food (g)"),
    "omega_3_dha": ("pufa_food", "cis n-3 C22:6 /100g food (g)"),
}

SFA_FALLBACK_COLUMNS = [
    "C4:0 /100g food (g)",
    "C6:0 /100g food (g)",
    "C8:0 /100g food (g)",
    "C10:0 /100g food (g)",
    "C12:0 /100g food (g)",
    "C13:0 /100g food (g)",
    "C14:0 /100g food (g)",
    "C15:0 /100g food (g)",
    "C16:0 /100g food (g)",
    "C17:0 /100g food (g)",
    "C18:0 /100g food (g)",
    "C19:0 /100g food (g)",
    "C20:0 /100g food (g)",
    "C22:0 /100g food (g)",
    "C24:0 /100g food (g)",
]

MUFA_FALLBACK_COLUMNS = [
    "cis C10:1 /100g food (g)",
    "cis C12:1 /100g food (g)",
    "cis C14:1 /100g food (g)",
    "cis C15:1 /100g food (g)",
    "cis C16:1 /100g food (g)",
    "cis C17:1 /100g food (g)",
    "cis C18:1 /100g food (g)",
    "cis C20:1 /100g food (g)",
    "cis C22:1 /100g food (g)",
    "cis C24:1 /100g food (g)",
]

PUFA_FALLBACK_COLUMNS = [
    "cis C16:2 /100g food (g)",
    "C16:3 /100g food (g)",
    "cis C16:4 /100g food (g)",
    "unknown C16 poly /100g food (g)",
    "cis n-6 C18:2 /100g food (g)",
    "cis n-3 C18:3 /100g food (g)",
    "cis n-6 C18:3 /100g food (g)",
    "cis n-3 C18:4 /100g food (g)",
    "unknown C18 poly /100g food (g)",
    "cis n-6 C20:2 /100g food (g)",
    "cis n-6 C20:3 /100g food (g)",
    "cis n-6 C20:4 /100g food (g)",
    "cis n-3 C20:5 /100g food (g)",
    "unknown C20 poly /100g food (g)",
    "cis n-3 C21:5 /100g food (g)",
    "cis n-6 C22:2 /100g food (g)",
    "cis n-6 C22:3 /100g food (g)",
    "cis n-6 C22:4 /100g food (g)",
    "cis n-3 C22:5 /100g food (g)",
    "cis n-3 C22:6 /100g food (g)",
    "unknown C22 poly /100g food (g)",
]

OMEGA3_FALLBACK_COLUMNS = [
    "cis n-3 C18:3 /100g food (g)",
    "cis n-3 C18:4 /100g food (g)",
    "cis n-3 C20:5 /100g food (g)",
    "cis n-3 C21:5 /100g food (g)",
    "cis n-3 C22:5 /100g food (g)",
    "cis n-3 C22:6 /100g food (g)",
]

OMEGA6_FALLBACK_COLUMNS = [
    "cis n-6 C18:2 /100g food (g)",
    "cis n-6 C18:3 /100g food (g)",
    "cis n-6 C20:2 /100g food (g)",
    "cis n-6 C20:3 /100g food (g)",
    "cis n-6 C20:4 /100g food (g)",
    "cis n-6 C22:2 /100g food (g)",
    "cis n-6 C22:3 /100g food (g)",
    "cis n-6 C22:4 /100g food (g)",
]

TRANS_FALLBACK_COLUMNS = {
    "mufa_food": ["trans monounsaturated /100g food (g)"],
    "pufa_food": ["trans poly /100g food (g)"],
}

PARENTHESIZED_NUMBER = re.compile(r"^\(([-+]?\d+(?:\.\d+)?)\)$")


def normalize_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None

        lowered = candidate.casefold()
        if lowered in {"n", "na", "n/a"}:
            return None
        if lowered in {"tr", "trace"}:
            return 0.0

        match = PARENTHESIZED_NUMBER.fullmatch(candidate)
        if match is not None:
            return float(match.group(1))

        numeric_candidate = candidate.replace(",", "")
        try:
            return float(numeric_candidate)
        except ValueError:
            return candidate
    return value


def canonicalize_header(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.replace("\n", " ").strip().split()).casefold()


def resolve_source_columns(
    headers: list[Any], source_columns: list[str]
) -> dict[str, str | None]:
    header_values = [h for h in headers if isinstance(h, str)]
    headers_by_canonical = {canonicalize_header(h): h for h in header_values}

    resolved: dict[str, str | None] = {}
    for source_column in source_columns:
        if source_column in header_values:
            resolved[source_column] = source_column
            continue

        source_canonical = canonicalize_header(source_column)
        direct_match = headers_by_canonical.get(source_canonical)
        if direct_match is not None:
            resolved[source_column] = direct_match
            continue

        startswith_match = next(
            (
                header
                for header in header_values
                if canonicalize_header(header).startswith(source_canonical)
            ),
            None,
        )
        resolved[source_column] = startswith_match

    return resolved


def load_sheet_rows(
    workbook_path: Path, sheet_name: str
) -> tuple[list[Any], dict[str, dict[str, Any]], list[str]]:
    wb = load_workbook(filename=workbook_path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if not headers:
        return headers, {}, []

    if canonicalize_header(headers[0]) == "":
        headers[0] = "Food Code"

    rows_by_code: dict[str, dict[str, Any]] = {}
    ordered_codes: list[str] = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_row = dict(zip(headers, row))
        code_raw = raw_row.get(headers[0])
        if not isinstance(code_raw, str):
            continue
        code = code_raw.strip()
        if not code:
            continue
        if code not in rows_by_code:
            ordered_codes.append(code)
        rows_by_code[code] = raw_row

    return headers, rows_by_code, ordered_codes


def get_sheet_value(
    sheet_rows: dict[str, dict[str, dict[str, Any]]],
    resolved_columns: dict[str, dict[str, str | None]],
    sheet_key: str,
    source_id: str,
    source_column: str,
) -> Any:
    row = sheet_rows[sheet_key].get(source_id)
    if row is None:
        return None

    resolved_column = resolved_columns[sheet_key].get(source_column)
    if resolved_column is None:
        return None

    return normalize_value(row.get(resolved_column))


def sum_sheet_values(
    sheet_rows: dict[str, dict[str, dict[str, Any]]],
    resolved_columns: dict[str, dict[str, str | None]],
    sheet_key: str,
    source_id: str,
    source_columns: list[str],
) -> float | None:
    total = 0.0
    found_value = False

    for source_column in source_columns:
        value = get_sheet_value(
            sheet_rows, resolved_columns, sheet_key, source_id, source_column
        )
        if isinstance(value, (int, float)):
            total += float(value)
            found_value = True

    return total if found_value else None


def sum_multi_sheet_values(
    sheet_rows: dict[str, dict[str, dict[str, Any]]],
    resolved_columns: dict[str, dict[str, str | None]],
    source_id: str,
    sheet_to_columns: dict[str, list[str]],
) -> float | None:
    total = 0.0
    found_value = False

    for sheet_key, source_columns in sheet_to_columns.items():
        value = sum_sheet_values(
            sheet_rows, resolved_columns, sheet_key, source_id, source_columns
        )
        if isinstance(value, (int, float)):
            total += float(value)
            found_value = True

    return total if found_value else None


def _resolve_source_literal(field_specs: dict[str, FieldSpec]) -> str:
    source_spec = field_specs.get("source")
    if (
        source_spec is not None
        and source_spec.kind == FieldKind.LITERAL
        and isinstance(source_spec.source, str)
        and source_spec.source
    ):
        return source_spec.source
    return "cofid"


def build_source_units_by_field(
    resolved_columns: dict[str, dict[str, str | None]]
) -> dict[str, str | None]:
    source_units: dict[str, str | None] = {}
    for target_field, (sheet_key, source_column) in DIRECT_FIELD_MAP.items():
        resolved_column = resolved_columns[sheet_key].get(source_column) or source_column
        source_units[target_field] = extract_unit_from_label(resolved_column)
    return source_units


def iter_rows(
    workbook_path: Path,
    core_fields: list[str] = CORE_FOOD_FIELDS,
    field_specs: dict[str, FieldSpec] = COFID_FIELD_SPECS,
):
    source_literal = _resolve_source_literal(field_specs)
    sheet_headers: dict[str, list[Any]] = {}
    sheet_rows: dict[str, dict[str, dict[str, Any]]] = {}
    source_ids: list[str] = []

    for sheet_key, sheet_name in COFID_SHEETS.items():
        headers, rows_by_code, ordered_codes = load_sheet_rows(
            workbook_path, sheet_name
        )
        sheet_headers[sheet_key] = headers
        sheet_rows[sheet_key] = rows_by_code
        if sheet_key == "proximates":
            source_ids = ordered_codes

    required_columns_by_sheet: dict[str, set[str]] = {
        key: set() for key in COFID_SHEETS
    }

    for sheet_key, source_column in DIRECT_FIELD_MAP.values():
        required_columns_by_sheet[sheet_key].add(source_column)

    for source_column in SFA_FALLBACK_COLUMNS:
        required_columns_by_sheet["sfa_food"].add(source_column)
    for source_column in MUFA_FALLBACK_COLUMNS:
        required_columns_by_sheet["mufa_food"].add(source_column)
    for source_column in PUFA_FALLBACK_COLUMNS:
        required_columns_by_sheet["pufa_food"].add(source_column)
    for source_column in OMEGA3_FALLBACK_COLUMNS:
        required_columns_by_sheet["pufa_food"].add(source_column)
    for source_column in OMEGA6_FALLBACK_COLUMNS:
        required_columns_by_sheet["pufa_food"].add(source_column)
    for sheet_key, source_columns in TRANS_FALLBACK_COLUMNS.items():
        for source_column in source_columns:
            required_columns_by_sheet[sheet_key].add(source_column)

    resolved_columns: dict[str, dict[str, str | None]] = {}
    for sheet_key, headers in sheet_headers.items():
        required_columns = sorted(required_columns_by_sheet[sheet_key])
        resolved_columns[sheet_key] = resolve_source_columns(headers, required_columns)
    source_units_by_field = build_source_units_by_field(resolved_columns)

    for source_id in source_ids:
        normalized: dict[str, Any] = {field: None for field in core_fields}
        normalized["source"] = source_literal
        normalized["portions"] = None

        for target_field, (sheet_key, source_column) in DIRECT_FIELD_MAP.items():
            value = get_sheet_value(
                sheet_rows, resolved_columns, sheet_key, source_id, source_column
            )
            if isinstance(value, (int, float)):
                value = convert_numeric_value(
                    value,
                    source_units_by_field.get(target_field),
                    CORE_FIELD_UNITS.get(target_field),
                    field=target_field,
                )
            normalized[target_field] = value

        if normalized.get("saturated_fats") is None:
            saturated_value = sum_sheet_values(
                sheet_rows,
                resolved_columns,
                "sfa_food",
                source_id,
                SFA_FALLBACK_COLUMNS,
            )
            normalized["saturated_fats"] = convert_numeric_value(
                saturated_value,
                "g",
                CORE_FIELD_UNITS.get("saturated_fats"),
                field="saturated_fats",
            )

        if normalized.get("monounsaturated_fat") is None:
            mufa_value = sum_sheet_values(
                sheet_rows,
                resolved_columns,
                "mufa_food",
                source_id,
                MUFA_FALLBACK_COLUMNS,
            )
            normalized["monounsaturated_fat"] = convert_numeric_value(
                mufa_value,
                "g",
                CORE_FIELD_UNITS.get("monounsaturated_fat"),
                field="monounsaturated_fat",
            )

        if normalized.get("polyunsaturated_fat") is None:
            pufa_value = sum_sheet_values(
                sheet_rows,
                resolved_columns,
                "pufa_food",
                source_id,
                PUFA_FALLBACK_COLUMNS,
            )
            normalized["polyunsaturated_fat"] = convert_numeric_value(
                pufa_value,
                "g",
                CORE_FIELD_UNITS.get("polyunsaturated_fat"),
                field="polyunsaturated_fat",
            )

        if normalized.get("omega_3_fats") is None:
            omega_3_value = sum_sheet_values(
                sheet_rows,
                resolved_columns,
                "pufa_food",
                source_id,
                OMEGA3_FALLBACK_COLUMNS,
            )
            normalized["omega_3_fats"] = convert_numeric_value(
                omega_3_value,
                "g",
                CORE_FIELD_UNITS.get("omega_3_fats"),
                field="omega_3_fats",
            )

        if normalized.get("omega_6_fats") is None:
            omega_6_value = sum_sheet_values(
                sheet_rows,
                resolved_columns,
                "pufa_food",
                source_id,
                OMEGA6_FALLBACK_COLUMNS,
            )
            normalized["omega_6_fats"] = convert_numeric_value(
                omega_6_value,
                "g",
                CORE_FIELD_UNITS.get("omega_6_fats"),
                field="omega_6_fats",
            )

        if normalized.get("trans_fats") is None:
            trans_value = sum_multi_sheet_values(
                sheet_rows, resolved_columns, source_id, TRANS_FALLBACK_COLUMNS
            )
            normalized["trans_fats"] = convert_numeric_value(
                trans_value,
                "g",
                CORE_FIELD_UNITS.get("trans_fats"),
                field="trans_fats",
            )

        if normalized.get("source_id") is None and normalized.get("name") is None:
            continue
        yield normalized


def register_subparser(subparsers):
    parser = subparsers.add_parser("cofid", help="Parse UK CoFID workbook")
    parser.add_argument("--workbook", required=True, help="Path to CoFID workbook (.xlsx)")
    parser.add_argument(
        "--output",
        default=None,
        help="Output JSONL path (default: data/outputs/cofid.jsonl, use '-' for stdout)",
    )
    parser.set_defaults(handler=run_from_args)


def run_from_args(args) -> None:
    output_path = resolve_output_path(args.output, SOURCE_NAME)
    write_jsonl(iter_rows(Path(args.workbook), CORE_FOOD_FIELDS), output_path)


def main(argv: list[str] | None = None):
    parser = argparse.ArgumentParser(description="CoFID parser")
    parser.add_argument("--workbook", required=True, help="Path to CoFID workbook (.xlsx)")
    parser.add_argument(
        "--output",
        default=None,
        help="Output JSONL path (default: data/outputs/cofid.jsonl, use '-' for stdout)",
    )
    args = parser.parse_args(argv)
    run_from_args(args)


if __name__ == "__main__":
    main()
