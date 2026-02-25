from __future__ import annotations

from pathlib import Path
import unittest

from openpyxl import load_workbook

from common.units import convert_numeric_value
from nutrient_mapping import (
    AUSTRALIA_FIELD_SPECS,
    CNF_FIELD_SPECS,
    CORE_FIELD_UNITS,
    NEVO_FIELD_SPECS,
    NEW_ZEALAND_FIELD_SPECS,
)
from parsers import australia, canada, cofid, nevo, new_zealand


class SourceUnitContractIntegrationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.repo_root = Path(__file__).resolve().parents[1]
        cls.inputs_dir = cls.repo_root / "data" / "inputs"

    def _require_inputs_root(self):
        if not self.inputs_dir.exists():
            self.skipTest(f"Integration inputs missing: {self.inputs_dir}")

    def _assert_convertible(
        self,
        source_name: str,
        field: str,
        source_unit: str | None,
        target_unit: str | None,
        source_ref: str,
    ) -> None:
        if target_unit is None:
            return
        try:
            convert_numeric_value(1.0, source_unit, target_unit, field=field)
        except Exception as exc:  # pragma: no cover - test assertion path
            self.fail(
                f"{source_name}: field={field} ref={source_ref} unit conversion "
                f"{source_unit!r}->{target_unit!r} failed: {exc}"
            )

    def test_australia_unit_contract(self):
        self._require_inputs_root()
        workbook_path = self.inputs_dir / "australian-food-composition-database.xlsx"
        if not workbook_path.exists():
            self.skipTest(f"Missing input file: {workbook_path}")

        wb = load_workbook(filename=workbook_path, data_only=True, read_only=True)
        ws = wb[australia.DEFAULT_SHEET_NAME]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
        source_map = australia.specs_to_source_map(AUSTRALIA_FIELD_SPECS)
        resolved = australia.resolve_source_columns(headers, source_map)
        source_units = australia.build_source_units(resolved)

        for field, target_unit in CORE_FIELD_UNITS.items():
            source_column = resolved.get(field)
            if source_column in ("", None):
                continue
            self._assert_convertible(
                "australia",
                field,
                source_units.get(field),
                target_unit,
                str(source_column),
            )

    def test_new_zealand_unit_contract(self):
        self._require_inputs_root()
        workbook_path = self.inputs_dir / "new-zealand-food-concise.xlsx"
        if not workbook_path.exists():
            self.skipTest(f"Missing input file: {workbook_path}")

        wb = load_workbook(filename=workbook_path, data_only=True, read_only=True)
        ws = wb[new_zealand.DEFAULT_SHEET_NAME]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        units_row = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
        source_map = new_zealand.specs_to_source_map(NEW_ZEALAND_FIELD_SPECS)
        resolved = new_zealand.resolve_source_columns(headers, source_map)
        units_by_header = new_zealand.build_unit_by_header(headers, units_row)
        source_units = new_zealand.build_source_units(resolved, units_by_header)

        for field, target_unit in CORE_FIELD_UNITS.items():
            source_column = resolved.get(field)
            if source_column in ("", None):
                continue
            self._assert_convertible(
                "new_zealand",
                field,
                source_units.get(field),
                target_unit,
                str(source_column),
            )

    def test_nevo_unit_contract(self):
        self._require_inputs_root()
        workbook_path = self.inputs_dir / "dutch-nutrient-database" / "NEVO2025_v9.0.xlsx"
        if not workbook_path.exists():
            self.skipTest(f"Missing input file: {workbook_path}")

        mapping = nevo.specs_to_source_map(NEVO_FIELD_SPECS)
        nutrient_metadata = nevo.get_nutrient_metadata(
            workbook_path, nevo.DEFAULT_NUTRIENTS_SHEET_NAME
        )
        source_units = nevo.build_source_units_by_field(mapping, nutrient_metadata)

        for field, target_unit in CORE_FIELD_UNITS.items():
            source_column = mapping.get(field)
            if source_column in ("", None):
                continue
            self._assert_convertible(
                "nevo",
                field,
                source_units.get(field),
                target_unit,
                str(source_column),
            )

    def test_canada_unit_contract(self):
        self._require_inputs_root()
        nutrient_name_path = self.inputs_dir / "canadian-nutrient-files" / "NUTRIENT NAME.csv"
        if not nutrient_name_path.exists():
            self.skipTest(f"Missing input file: {nutrient_name_path}")

        mapping = canada.specs_to_symbol_map(CNF_FIELD_SPECS)
        _, nutrient_by_symbol = canada.load_nutrient_symbols(nutrient_name_path)
        unit_by_symbol = {
            symbol: (
                row.get("NutrientUnit").strip()
                if isinstance(row.get("NutrientUnit"), str)
                else None
            )
            for symbol, row in nutrient_by_symbol.items()
        }

        for field, target_unit in CORE_FIELD_UNITS.items():
            source_symbol = mapping.get(field)
            if source_symbol in ("", None):
                continue
            self._assert_convertible(
                "canada",
                field,
                unit_by_symbol.get(source_symbol),
                target_unit,
                str(source_symbol),
            )

    def test_cofid_unit_contract(self):
        self._require_inputs_root()
        workbook_path = self.inputs_dir / "CoFID.xlsx"
        if not workbook_path.exists():
            self.skipTest(f"Missing input file: {workbook_path}")

        sheet_headers: dict[str, list] = {}
        for sheet_key, sheet_name in cofid.COFID_SHEETS.items():
            headers, _, _ = cofid.load_sheet_rows(workbook_path, sheet_name)
            sheet_headers[sheet_key] = headers

        required_columns_by_sheet: dict[str, set[str]] = {
            key: set() for key in cofid.COFID_SHEETS
        }
        for sheet_key, source_column in cofid.DIRECT_FIELD_MAP.values():
            required_columns_by_sheet[sheet_key].add(source_column)
        for source_column in cofid.SFA_FALLBACK_COLUMNS:
            required_columns_by_sheet["sfa_food"].add(source_column)
        for source_column in cofid.MUFA_FALLBACK_COLUMNS:
            required_columns_by_sheet["mufa_food"].add(source_column)
        for source_column in cofid.PUFA_FALLBACK_COLUMNS:
            required_columns_by_sheet["pufa_food"].add(source_column)
        for source_column in cofid.OMEGA3_FALLBACK_COLUMNS:
            required_columns_by_sheet["pufa_food"].add(source_column)
        for source_column in cofid.OMEGA6_FALLBACK_COLUMNS:
            required_columns_by_sheet["pufa_food"].add(source_column)
        for sheet_key, source_columns in cofid.TRANS_FALLBACK_COLUMNS.items():
            for source_column in source_columns:
                required_columns_by_sheet[sheet_key].add(source_column)

        resolved_columns: dict[str, dict[str, str | None]] = {}
        for sheet_key, headers in sheet_headers.items():
            resolved_columns[sheet_key] = cofid.resolve_source_columns(
                headers, sorted(required_columns_by_sheet[sheet_key])
            )
        source_units = cofid.build_source_units_by_field(resolved_columns)

        for field, target_unit in CORE_FIELD_UNITS.items():
            if field in cofid.DIRECT_FIELD_MAP:
                self._assert_convertible(
                    "cofid",
                    field,
                    source_units.get(field),
                    target_unit,
                    field,
                )

        # CoFID fallback fields are sums of fatty-acid subcomponents expressed in grams.
        fallback_fields = [
            "saturated_fats",
            "monounsaturated_fat",
            "polyunsaturated_fat",
            "omega_3_fats",
            "omega_6_fats",
            "trans_fats",
        ]
        for field in fallback_fields:
            self._assert_convertible("cofid", field, "g", CORE_FIELD_UNITS.get(field), "fallback")


if __name__ == "__main__":
    unittest.main()
